from openpyxl import load_workbook
import datetime
import time


class UtilityPackage():
    def padd_zeroes(self, data):
        return str(data).zfill(6)

    def date_picker_data_strptime(self, date_input_value):
        date_input_value_to_string = datetime.datetime.strptime(date_input_value, "%m/%d/%Y")
        date_picker_year = date_input_value_to_string.strftime("%Y")
        date_picker_month = date_input_value_to_string.strftime("%m").lstrip("0")
        date_picker_month_txt = date_input_value_to_string.strftime("%B").lstrip("0")
        date_picker_date = date_input_value_to_string.strftime("%d").lstrip("0")
        formatted_date_text = f"{date_picker_month_txt} {date_picker_date}, {date_picker_year}"

        return date_picker_year, date_picker_month, date_picker_date, formatted_date_text

    def month_number_conversion_xxx(self, date_input_value):
        match date_input_value:
            case "1":
                return "Jan"
            case "2":
                return "Feb"
            case "3":
                return "Mar"
            case "4":
                return "Apr"
            case "5":
                return "May"
            case "6":
                return "Jun"
            case "7":
                return "Jul"
            case "8":
                return "Aug"
            case "9":
                return "Sep"
            case "10":
                return "Oct"
            case "11":
                return "Nov"
            case "12":
                return "Dec"

    def cbu_excel(self, data_path, data_entry_num, data_ecp_information):

        open_excel_path_location = data_path
        get_work_book = load_workbook(open_excel_path_location)
        get_active_sheet = get_work_book.active
        get_total_rows = get_active_sheet.max_row

        for x in range(get_total_rows):
            if x > 0:
                CellRow = x + 1
                extension_num = str(x).zfill(4)

                engine_data = f'{data_entry_num["cbu_engine"]}{data_ecp_information}{extension_num}'
                chassis_data = f'{data_entry_num["cbu_chassis"]}{data_ecp_information}{extension_num}'
                vin_data = f'{data_entry_num["cbu_vin"]}{data_ecp_information}{extension_num}'

                if get_active_sheet.cell(row=CellRow, column=9).value != "":
                    get_active_sheet.cell(row=CellRow, column=9, value=engine_data)
                if get_active_sheet.cell(row=CellRow, column=10).value != "":
                    get_active_sheet.cell(row=CellRow, column=10, value=chassis_data)
                if get_active_sheet.cell(row=CellRow, column=11).value != "":
                    get_active_sheet.cell(row=CellRow, column=11, value=vin_data)

        get_work_book.save(open_excel_path_location)

